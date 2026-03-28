import sqlite3
from pathlib import Path

import openpyxl


# Adjust these paths if needed.
DB_PATH = Path("output/yonsei_medicine_faculty.db")
XLSX_PATH = Path("output/yonsei_medicine_faculty.xlsx")


def fix_db(db_path: Path):
    if not db_path.exists():
        print(f"[DB] not found: {db_path}")
        return

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Reset obviously polluted fields.
    cur.execute("""
        UPDATE faculty
        SET
            name_en = CASE
                WHEN name_en IN ('"', 'Yonsei Faculty Information', 'Department', 'Name')
                  OR name_en LIKE '%Campus%'
                  OR name_en LIKE '%Department of%'
                THEN ''
                ELSE name_en
            END,
            title_ko = CASE
                WHEN lower(title_ko) LIKE '%@%'
                  OR title_ko IN ('Yonsei Faculty Information', 'Department', 'Name', '논문', '학술활동')
                THEN ''
                ELSE title_ko
            END,
            office = CASE
                WHEN office IN ('논문', '학술활동', 'Yonsei Faculty Information')
                THEN ''
                ELSE office
            END
    """)

    # For known Family Medicine records from the current crawl results.
    cur.execute("""
        UPDATE faculty
        SET name_ko='심재용', name_en='', title_ko='', office=''
        WHERE detail_url LIKE '%userId=%2By5TOOv1KyAQ4AQ%2Fkw5CXQ%3D%3D%'
    """)
    cur.execute("""
        UPDATE faculty
        SET name_ko='이용제', name_en='', title_ko='', office=''
        WHERE detail_url LIKE '%userId=6KnJVpJaX0vzCylhle93gA%3D%3D%'
    """)
    cur.execute("""
        UPDATE faculty
        SET name_ko='이지원', name_en='', title_ko='', office=''
        WHERE detail_url LIKE '%userId=VOfDxCs3oWnDxbHnHDypRg%3D%3D%'
    """)

    conn.commit()
    conn.close()
    print(f"[DB] corrected: {db_path}")


def find_header_map(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            headers[str(value).strip()] = col
    return headers


def fix_xlsx(xlsx_path: Path):
    if not xlsx_path.exists():
        print(f"[XLSX] not found: {xlsx_path}")
        return

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    headers = find_header_map(ws)

    required = ["name_ko", "name_en", "title_ko", "office", "detail_url"]
    missing = [h for h in required if h not in headers]
    if missing:
        print(f"[XLSX] missing columns: {missing}")
        return

    c_name_ko = headers["name_ko"]
    c_name_en = headers["name_en"]
    c_title_ko = headers["title_ko"]
    c_office = headers["office"]
    c_detail_url = headers["detail_url"]

    for row in range(2, ws.max_row + 1):
        detail_url = str(ws.cell(row, c_detail_url).value or "").strip()
        name_en = str(ws.cell(row, c_name_en).value or "").strip()
        title_ko = str(ws.cell(row, c_title_ko).value or "").strip()
        office = str(ws.cell(row, c_office).value or "").strip()

        if (
            name_en in {'"', 'Yonsei Faculty Information', 'Department', 'Name'}
            or 'Campus' in name_en
            or 'Department of' in name_en
        ):
            ws.cell(row, c_name_en).value = ""

        if (
            "@" in title_ko
            or title_ko in {'Yonsei Faculty Information', 'Department', 'Name', '논문', '학술활동'}
        ):
            ws.cell(row, c_title_ko).value = ""

        if office in {'논문', '학술활동', 'Yonsei Faculty Information'}:
            ws.cell(row, c_office).value = ""

        if "%2By5TOOv1KyAQ4AQ%2Fkw5CXQ%3D%3D" in detail_url:
            ws.cell(row, c_name_ko).value = "심재용"
            ws.cell(row, c_name_en).value = ""
            ws.cell(row, c_title_ko).value = ""
            ws.cell(row, c_office).value = ""

        elif "6KnJVpJaX0vzCylhle93gA%3D%3D" in detail_url:
            ws.cell(row, c_name_ko).value = "이용제"
            ws.cell(row, c_name_en).value = ""
            ws.cell(row, c_title_ko).value = ""
            ws.cell(row, c_office).value = ""

        elif "VOfDxCs3oWnDxbHnHDypRg%3D%3D" in detail_url:
            ws.cell(row, c_name_ko).value = "이지원"
            ws.cell(row, c_name_en).value = ""
            ws.cell(row, c_title_ko).value = ""
            ws.cell(row, c_office).value = ""

    corrected_path = xlsx_path.with_name(xlsx_path.stem + "_fixed.xlsx")
    wb.save(corrected_path)
    print(f"[XLSX] corrected: {corrected_path}")


if __name__ == "__main__":
    fix_db(DB_PATH)
    fix_xlsx(XLSX_PATH)
